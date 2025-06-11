from fastapi import FastAPI, UploadFile, File, HTTPException, Request
from fastapi.responses import JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Dict
import pandas as pd
import fitz  # PyMuPDF
import re
from openpyxl import load_workbook

app = FastAPI()

# Permitir CORS desde cualquier origen
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

class ConsumoRequest(BaseModel):
    dias_factura: int
    potencia: Dict[str, float]
    energia: Dict[str, float]

@app.get("/ping")
async def ping():
    return {"pong": True}

def extract(pattern: str, text: str, fmt=float, label: str = None, default=None):
    """Busca pattern en text y devuelve group(1) formateado o default si no encuentra."""
    m = re.search(pattern, text, flags=re.IGNORECASE)
    if not m:
        if default is not None:
            return default
        lbl = f" '{label}'" if label else ""
        raise ValueError(f"No se encontró el campo{lbl}.")
    val = m.group(1).replace(",", ".")
    return fmt(val)

@app.post("/analizar-factura")
async def analizar_factura(file: UploadFile = File(...)):
    try:
        contents = await file.read()
        doc = fitz.open(stream=contents, filetype="pdf")
        text = "".join(page.get_text() for page in doc)
        doc.close()

        dias             = extract(r"DIAS FACTURADOS:\s*(\d+)", text, int, "Días")
        potencia_punta   = extract(r"Potencia punta:\s*([\d,]+)\s*kW", text, float, "Potencia punta")
        potencia_valle   = extract(r"Potencia valle:\s*([\d,]+)\s*kW", text, float, "Potencia valle")
        consumo_punta    = extract(r"punta:\s*([\d,]+)\s*kWh", text, float, "Consumo punta")
        consumo_llano    = extract(r"llano:\s*([\d,]+)\s*kWh", text, float, "Consumo llano")
        consumo_valle    = extract(r"valle[: ]\s*([\d,]+)\s*kWh", text, float, "Consumo valle")

        total_factura    = extract(r"TOTAL IMPORTE FACTURA\D*([\d,]+,[\d]{2})\s*€", text, float, "Total factura")
        factura_impuesto = extract(r"IVA.*?([\d,]+,[\d]{2})\s*€", text, float, "IVA", default=0.0)
        factura_alquiler = extract(r"Alquiler equipos medida.*?([\d,]+,[\d]{2})\s*€", text, float, "Alquiler", default=0.0)

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al extraer datos del PDF: {e}")

    return {
        "dias_factura": dias,
        "potencia": {"punta": potencia_punta, "valle": potencia_valle},
        "energia": {"punta": consumo_punta, "llano": consumo_llano, "valle": consumo_valle},
        "factura_total": round(total_factura, 2),
        "factura_impuesto": round(factura_impuesto, 2),
        "factura_alquiler": round(factura_alquiler, 2)
    }

@app.post("/comparar-tarifas")
async def comparar_tarifas(consumo: ConsumoRequest):
    try:
        excel = "Comparador electricidad.v3 (1).xlsx"
        df = pd.read_excel(excel, sheet_name="Comparador")

        # Precios de potencia
        pot = df.iloc[[0,6,7],4:].transpose().reset_index(drop=True)
        pot.columns = ["nombre","potencia_punta","potencia_valle"]
        pot = pot.apply(pd.to_numeric, errors="coerce")

        # Precios de energía
        ene = df.iloc[[11,12,13],4:].transpose().reset_index(drop=True)
        ene.columns = ["energia_punta","energia_llano","energia_valle"]
        ene = ene.apply(pd.to_numeric, errors="coerce")

        # Enlaces (fila 2)
        wb = load_workbook(excel, read_only=True)
        ws = wb["Comparador"]
        enlaces = [cell.hyperlink.target if cell.hyperlink else "" for cell in ws[2][4:]]
        wb.close()

        tarifas = pd.concat([pot, ene], axis=1).reset_index(drop=True)
        tarifas["enlace"] = enlaces
        tarifas = tarifas.dropna(subset=["potencia_punta","potencia_valle","energia_punta"])

        resultados = []
        for _, r in tarifas.iterrows():
            cp = (
                consumo.potencia["punta"]*r["potencia_punta"]*consumo.dias_factura +
                consumo.potencia["valle"]*r["potencia_valle"]*consumo.dias_factura
            )
            ce = (
                consumo.energia["punta"]*r["energia_punta"] +
                consumo.energia["llano"]*r["energia_llano"] +
                consumo.energia["valle"]*r["energia_valle"]
            )
            var = round(cp + ce, 2)
            # Incluimos aquí el coste fijo extraído antes
            fijo = consumo.__dict__.get("factura_impuesto", 0) + consumo.__dict__.get("factura_alquiler", 0)
            total = round(var + fijo, 2)
            resultados.append({
                "tarifa": r["nombre"],
                "coste_variable": var,
                "coste_fijo": round(fijo, 2),
                "coste_total": total,
                "enlace": r["enlace"]
            })

        resultados.sort(key=lambda x: x["coste_total"])
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al comparar tarifas: {e}")

    return JSONResponse(resultados)
